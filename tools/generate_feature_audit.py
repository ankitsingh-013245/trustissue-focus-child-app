from __future__ import annotations

from collections import Counter
from datetime import date
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


PROJECT_ROOT = Path(__file__).resolve().parents[1]
OUTPUT_PATH = PROJECT_ROOT / "TrustIssue_Feature_Audit_2026-07-23.xlsx"
AUDIT_DATE = date(2026, 7, 23)

STATUS_IMPLEMENTED = "Implemented — runtime validation pending"
STATUS_LIMITED = "Implemented with platform limits"
STATUS_UI_ONLY = "UI-only — Coming soon"
STATUS_REMOVED = "Removed by product decision"
STATUS_LEGACY = "Legacy/unreferenced — cleanup pending"


def feature(
    feature_id: str,
    area: str,
    name: str,
    status: str,
    behavior: str,
    mechanism: str,
    permissions: str,
    accessibility: str,
    reliability: int | None,
    limitations: str,
    verification: str,
    sources: str,
    priority: str = "P1",
    pending: str = "Run listed device tests; no compile/runtime validation was performed.",
) -> dict[str, object]:
    return {
        "ID": feature_id,
        "Area": area,
        "Feature": name,
        "Implementation status": status,
        "What the user sees / what it does": behavior,
        "How it works internally": mechanism,
        "Permissions / access": permissions,
        "Accessibility role": accessibility,
        "Expected reliability (1-5)": reliability,
        "Known limitations / risks": limitations,
        "How to verify / accept": verification,
        "Primary source files": sources,
        "Release priority": priority,
        "Runtime validation": "NOT RUN — user requested no compilation",
        "Pending action": pending,
    }


FEATURES = [
    feature(
        "F-001", "Onboarding", "Permission-free onboarding", STATUS_IMPLEMENTED,
        "Introduces Focus without forcing permission screens during first launch.",
        "The Flutter onboarding screen contains no permission requests; access is requested only when a dependent feature is started.",
        "None", "Not required", 5,
        "A future onboarding edit could accidentally add eager permission calls.",
        "Fresh-install test: complete onboarding and confirm Android shows no Usage, Overlay, Accessibility, VPN, DND, notification, or battery prompt.",
        "lib/screens/onboarding_screen.dart; lib/screens/home_shell.dart",
    ),
    feature(
        "F-002", "Onboarding", "Default PDF reader choice", STATUS_IMPLEMENTED,
        "Lets the user choose a PDF reader for study handoffs.",
        "Native package discovery finds eligible PDF handlers; the selected package is persisted locally and checked before App Study starts.",
        "Package visibility queries only", "Not required for selection; automatic cross-app recognition is more reliable with Accessibility", 4,
        "If the selected reader is uninstalled or becomes a browser, the user must choose again.",
        "Install two PDF readers; choose one; restart app; start App Study; uninstall it and confirm a new selection is requested.",
        "lib/screens/onboarding_screen.dart; lib/screens/settings_screen.dart; android/.../StudyToolResolver.kt; android/.../MainActivity.kt",
    ),
    feature(
        "F-003", "Navigation & UI", "Four-tab navigation", STATUS_IMPLEMENTED,
        "Provides Focus, Blocks, Reports, and Settings tabs.",
        "Flutter HomeShell swaps four embedded screens and refreshes usage or analytics when relevant tabs open.",
        "None", "Not required", 5,
        "Small-screen and large-font layouts still need device checks.",
        "Open every tab repeatedly during idle and active Focus; verify state is preserved and there is no clipped navigation at 1.0x–2.0x font scale.",
        "lib/screens/home_shell.dart",
    ),
    feature(
        "F-004", "Focus UI", "Focus home dashboard", STATUS_IMPLEMENTED,
        "Shows usage time, focus time, selected apps, current mode, and a central start/edit control.",
        "Flutter combines locally stored configuration with native session state, break state, usage summary, and app icons.",
        "Usage Access is needed for real usage totals; UI itself needs none", "Not required", 4,
        "Usage totals show zero until Usage Access is granted; icon loading is asynchronous.",
        "Compare home totals with Android Digital Wellbeing for the same day; verify active session timer updates every second.",
        "lib/screens/focus_home_screen.dart; lib/screens/home_shell.dart; lib/services/native_bridge.dart",
    ),
    feature(
        "F-005", "Focus UI", "Circular start/edit interaction", STATUS_IMPLEMENTED,
        "Tapping the circle or Edit opens a compact Focus setup sheet; active Focus shows live state.",
        "Flutter modal sheet edits an in-memory StudySettings copy and persists only confirmed values.",
        "None", "Not required", 5,
        "Gesture and sheet-height behavior needs testing on small devices.",
        "Tap circle and Edit; change values; cancel and confirm; verify cancel does not persist and confirm does.",
        "lib/screens/focus_home_screen.dart",
    ),
    feature(
        "F-006", "Session type", "Timer session", STATUS_IMPLEMENTED,
        "Counts down for a selected fixed duration.",
        "Duration is capped by protection mode, persisted in Flutter preferences, mirrored into native state, and guarded by an Android timeout alarm.",
        "Usage Access + Display over apps when starting Focus; notification prompt is optional", "Optional performance enhancement", 4,
        "AlarmManager setAndAllowWhileIdle is not exact; the foreground observer also performs completion checks.",
        "Run 1-, 5-, and 30-minute timers with screen on/off; confirm end occurs once and protection is removed.",
        "lib/screens/focus_home_screen.dart; lib/services/settings_store.dart; android/.../TrackerConfig.kt; android/.../FocusSessionTimeoutReceiver.kt",
    ),
    feature(
        "F-007", "Session type", "Duration wheel", STATUS_IMPLEMENTED,
        "Allows hours/minutes selection within the selected mode's safety maximum.",
        "Flutter wheel clamps Normal to 24h, Strict to 8h, and Locked Strict to 3h.",
        "None", "Not required", 5,
        "Boundary values still require UI and persistence tests.",
        "Test 0, 1 minute, mode maximum, and one step beyond maximum; verify invalid values cannot start.",
        "lib/screens/focus_home_screen.dart; lib/services/settings_store.dart",
    ),
    feature(
        "F-008", "Session type", "Stopwatch session", STATUS_IMPLEMENTED,
        "Counts focused time upward until the user exits; earned breaks appear as focus time accumulates.",
        "Native state stores Stopwatch mode, uses elapsedRealtime for active spans, and applies a safety deadline. Locked Strict is automatically disabled.",
        "Same as Focus start; DND additionally for Strict Stopwatch", "Optional performance enhancement", 4,
        "Stopwatch is intentionally not available with Locked Strict; long-duration safety deadline must be tested.",
        "Run Normal and Strict Stopwatch; background/reopen app; reboot test; confirm count does not double and exit rules are correct.",
        "lib/screens/focus_home_screen.dart; lib/services/settings_store.dart; android/.../TrackerConfig.kt",
    ),
    feature(
        "F-009", "Session type", "Pomodoro", STATUS_REMOVED,
        "No Pomodoro tab or configuration is presented.",
        "Product decision: Timer and Stopwatch remain; automatic breaks replace a separate Pomodoro mode.",
        "None", "Not applicable", None,
        "Old design screenshots may still create expectation.",
        "Search UI and source references; confirm no active Pomodoro navigation or start path exists.",
        "lib/screens/focus_home_screen.dart",
        priority="N/A", pending="Keep removed unless product direction changes.",
    ),
    feature(
        "F-010", "Tracking", "Continuous timer (Book Study)", STATUS_IMPLEMENTED,
        "Focus time continues against a wall-clock session even when no selected study app is open, except during breaks.",
        "Native Book Study computes focused time from the session clock and pushes the deadline out by permitted break time.",
        "Focus permissions", "Optional performance enhancement", 4,
        "Wall-clock recovery after reboot depends on persisted anchors; clock-change receiver must be tested.",
        "Start 10-minute Continuous timer; use launcher/allowed apps; take break; verify focus and end time calculations.",
        "lib/screens/focus_home_screen.dart; android/.../TrackerConfig.kt",
    ),
    feature(
        "F-011", "Tracking", "Selected apps only (App Study)", STATUS_IMPLEMENTED,
        "Counts focus only while a selected study app or approved study tool is in use.",
        "Allowlist policy resumes and pauses native timing according to observed foreground package transitions.",
        "Usage Access + Display over apps", "Optional but materially improves transition speed and PDF/Gallery handoffs", 3,
        "UsageStats polling can miss very short transitions; PDF/Gallery intent origin cannot always be inferred without Accessibility.",
        "Switch among selected app, launcher, unselected app, PDF, and Gallery; compare elapsed focus with a stopwatch.",
        "lib/screens/focus_home_screen.dart; android/.../TrackerConfig.kt; android/.../SelfControlAccessibilityService.kt; android/.../FocusUsageMonitorService.kt",
    ),
    feature(
        "F-012", "App policy", "Block selected apps (blocklist)", STATUS_IMPLEMENTED,
        "Only chosen distracting apps are blocked during Focus; other non-safety apps remain usable.",
        "FocusPolicyEngine returns a block decision when the foreground package is in the persisted blocked set.",
        "Usage Access + Display over apps", "Optional speed enhancement", 4,
        "Foreground detection and background-start restrictions vary by Android/OEM.",
        "Select three apps; start Focus; test selected, unselected, safety, browser, and TrustIssue packages.",
        "lib/screens/select_apps_screen.dart; android/.../FocusPolicyEngine.kt; android/.../TrackerConfig.kt",
    ),
    feature(
        "F-013", "App policy", "Allow only selected apps (allowlist)", STATUS_IMPLEMENTED,
        "Selected study apps and safety apps are usable; other apps are blocked.",
        "FocusPolicyEngine permits packages in the allowlist plus safety/system exceptions.",
        "Usage Access + Display over apps; VPN consent if an allowed browser exists", "Optional speed enhancement", 4,
        "System helper classification and OEM packages require broad device testing.",
        "Allow one study app; test all other launcher apps, Settings, Phone, Messages, Maps, keyboard, file picker, PDF, and browser.",
        "lib/screens/select_apps_screen.dart; android/.../FocusPolicyEngine.kt; android/.../StudyToolResolver.kt",
    ),
    feature(
        "F-014", "App selection", "Installed app discovery", STATUS_IMPLEMENTED,
        "Lists launchable installed apps, excluding TrustIssue itself.",
        "Native PackageManager queries launcher activities, de-duplicates package names, and sorts by label.",
        "Package visibility query declarations", "Not required", 4,
        "Work-profile, Secure Folder, hidden, instant, or non-launcher components may not appear.",
        "Compare picker against launcher across Samsung/Xiaomi/Pixel and work-profile devices.",
        "android/.../MainActivity.kt; android/app/src/main/AndroidManifest.xml",
    ),
    feature(
        "F-015", "App selection", "Search, categories, select-all, and icons", STATUS_IMPLEMENTED,
        "Searches by app name/package, surfaces known categories first, supports bulk selection, and shows app icons.",
        "Flutter filters/sorts native app metadata; AppIconCache requests PNG icons over MethodChannel and caches them.",
        "None beyond app discovery", "Not required", 4,
        "Category map is curated and incomplete; large app inventories may need performance profiling.",
        "Test 300+ installed apps, rapid search, select-all on filtered results, icon cache after scroll/reopen.",
        "lib/screens/select_apps_screen.dart; lib/services/app_icon_cache.dart; lib/widgets/app_icon_tile.dart",
    ),
    feature(
        "F-016", "Safety", "Always-allowed safety apps", STATUS_IMPLEMENTED,
        "Phone, messages, settings, maps and essential system helpers stay reachable and appear read-only in selection.",
        "Native alwaysAllowedPackages and system/transient-window rules override normal block decisions.",
        "None", "Not required", 4,
        "OEM package names differ; package list needs device-matrix validation and must not over-allow a distracting app.",
        "During allowlist Focus test calls, SMS, Settings, emergency dialer, maps, keyboard, permission controller, chooser, and file picker.",
        "android/.../MainActivity.kt; android/.../FocusForegroundEventPolicy.kt; lib/screens/select_apps_screen.dart",
    ),
    feature(
        "F-017", "Protection mode", "Normal Focus", STATUS_IMPLEMENTED,
        "Up to 24 hours; user can end at any time after confirmation.",
        "Native observer enforces the selected app policy; Flutter stop flow calls native deactivation and preserves completed local metrics.",
        "Usage Access + Display over apps; notification optional; VPN only if browser allowed", "Optional speed enhancement", 4,
        "Compatible mode depends on foreground service survival and OEM battery policy.",
        "Start/end/restart sessions; revoke notification permission; lock screen; verify blocked apps and cleanup.",
        "lib/screens/home_shell.dart; android/.../TrackerConfig.kt; android/.../FocusUsageMonitorService.kt",
    ),
    feature(
        "F-018", "Protection mode", "Strict Focus", STATUS_IMPLEMENTED,
        "Up to 8 hours; enables DND and uses a protected emergency-exit process.",
        "The same blocking engine runs with strict flags; StrictFocusDndController applies alarms-only DND and FocusGateActivity owns emergency exit.",
        "Usage Access + Display over apps + DND policy access; notification optional; VPN conditionally", "Optional speed enhancement", 4,
        "DND behavior and visual suppression vary by Android/OEM; emergency calls must remain possible.",
        "Test 1-minute and maximum-boundary Strict sessions, incoming call/SMS/app notifications, DND restore, and emergency exit.",
        "lib/screens/strict_preparation_screen.dart; android/.../StrictFocusDndController.kt; android/.../FocusGateActivity.kt",
    ),
    feature(
        "F-019", "Protection mode", "Locked Strict Focus", STATUS_IMPLEMENTED,
        "Fixed Timer only, maximum 3 hours, DND active, and no emergency exit or early-stop control.",
        "Locked flag is persisted; native stop API and gate UI reject exit while the session is active; timeout completes it.",
        "Usage Access + Display over apps + DND; VPN conditionally", "Optional speed enhancement", 4,
        "Any bug in timeout/restoration has high user impact; power-cycle and clock-change tests are mandatory before release.",
        "Verify Stopwatch toggle disables Locked; try all UI/native exit paths; reboot/update/change time; confirm automatic end and DND release.",
        "lib/services/settings_store.dart; lib/screens/home_shell.dart; android/.../MainActivity.kt; android/.../TrackerConfig.kt",
        priority="P0",
    ),
    feature(
        "F-020", "Protection mode", "Strict preparation countdown", STATUS_IMPLEMENTED,
        "Shows a 60-second preparation screen with wait, skip, or cancel before Strict starts.",
        "Flutter full-screen countdown returns a boolean; native activation happens only after confirmation.",
        "None beyond permissions already checked", "Not required", 5,
        "App lifecycle interruption during the preparation screen needs testing.",
        "Wait, skip, cancel, background, rotate, and kill during preparation; confirm Focus starts only after explicit completion.",
        "lib/screens/strict_preparation_screen.dart; lib/screens/home_shell.dart",
    ),
    feature(
        "F-021", "Strict exit", "Protected emergency exit", STATUS_IMPLEMENTED,
        "Strict requires a 60-second wait and at least 10 words; cancel returns to Focus.",
        "Native FocusGateActivity validates word count and countdown, records attempt/cancel/success metrics, then deactivates Strict.",
        "None additional", "Not required", 4,
        "Typed reason is described as local report data; privacy and deletion behavior must match this promise.",
        "Attempt with 0/9/10 words; background/rotate; cancel; complete; verify DND and protection cleanup plus analytics.",
        "android/.../FocusGateActivity.kt; android/.../TrackerConfig.kt",
        priority="P0",
    ),
    feature(
        "F-022", "Strict exit", "Locked Strict has no emergency exit", STATUS_IMPLEMENTED,
        "No exit button is rendered, and API attempts return a Locked Strict message.",
        "Both Flutter and native layers check the locked flag; timeout is the intended completion route.",
        "None additional", "Not required", 4,
        "Must verify notification actions, recents, deep links, process restart, and settings changes cannot bypass it.",
        "Enumerate every stop path and attempt during active Locked Strict; verify all fail except natural timeout.",
        "lib/screens/home_shell.dart; android/.../MainActivity.kt; android/.../FocusGateActivity.kt; android/.../FocusPillController.kt",
        priority="P0",
    ),
    feature(
        "F-023", "Notifications", "Strict-mode notification blocking via DND", STATUS_IMPLEMENTED,
        "Strict and Locked Strict use alarms-only DND; Normal Focus does not block notifications.",
        "App-owned DND state is applied, previous user filter/policy is saved, and only app-owned state is released.",
        "DND policy access", "Not required", 4,
        "Calls, alarms, OEM overlays, companion devices, and target-35 DND rule behavior differ; user can revoke access mid-session.",
        "Test incoming calls, emergency call path, SMS, heads-up, lock-screen notifications, alarms, and restore of pre-existing DND.",
        "android/.../StrictFocusDndController.kt; lib/screens/home_shell.dart",
        priority="P0",
    ),
    feature(
        "F-024", "Breaks", "Automatic break earning", STATUS_IMPLEMENTED,
        "Timer earns one break per 30 configured minutes; Stopwatch earns breaks from completed focused time.",
        "Native TrackerConfig calculates allowed/used break budget and exposes the next available break duration.",
        "Focus permissions", "Not required for calculation; observer is needed for enforcement", 4,
        "Boundary timing and interrupted sessions need tests; UI wording must explain the rule.",
        "Test 29:59, 30:00, 59:59, 60:00, Stopwatch pause/resume, and app restarts.",
        "lib/services/settings_store.dart; android/.../TrackerConfig.kt",
    ),
    feature(
        "F-025", "Breaks", "Break app picker", STATUS_IMPLEMENTED,
        "Lets the user select up to three apps for a temporary break.",
        "FocusGateActivity presents eligible app options and persists the selected break package(s) with a timeout.",
        "Display over apps; foreground observer", "Optional speed enhancement", 4,
        "The user-facing copy says up to three; all multi-selection persistence/enforcement combinations need runtime confirmation.",
        "Select 0/1/3/4 apps; start break; verify only selected apps and safety apps open.",
        "android/.../FocusGateActivity.kt; android/.../TrackerConfig.kt",
    ),
    feature(
        "F-026", "Breaks", "Five-minute break slices and early return", STATUS_IMPLEMENTED,
        "Break time is temporary; ending early returns unused time and restores Focus rules.",
        "Elapsed and wall-clock anchors track actual/allotted/unused break time; manual end and alarm expiry use one cleanup path.",
        "None additional", "Optional speed enhancement", 4,
        "Doze can delay an inexact alarm, so foreground reconciliation must catch expiry immediately on next observation.",
        "End at 30s and let full break expire with screen on/off; verify remaining budget, app blocking, timer, and reports.",
        "android/.../TrackerConfig.kt; android/.../FocusBreakTimeoutReceiver.kt; android/.../FocusGateActivity.kt",
    ),
    feature(
        "F-027", "Floating controls", "Floating Focus pill", STATUS_IMPLEMENTED,
        "Shows a draggable live timer over the active study/break app; expands to break, exit, and collapse controls.",
        "Overlay WindowManager view persists position; visibility policy hides it outside the relevant scope and auto-collapses expanded controls.",
        "Display over apps", "Works in both engines", 4,
        "OEM overlay restrictions, orientation, cutouts, multi-window, and accidental obstruction need testing.",
        "Drag to all edges, rotate, collapse/expand, long-press actions, change apps, lock screen, and restart service.",
        "android/.../FocusPillController.kt; android/.../FocusPillVisibilityPolicy.kt",
    ),
    feature(
        "F-028", "Blocking engine", "Usage Access compatible observer", STATUS_LIMITED,
        "Provides app blocking without Accessibility by detecting foreground package changes and opening the block gate.",
        "A foreground service polls UsageStats roughly every 300ms while active and slows when idle; overlay safety UI appears before the richer Activity.",
        "Usage Access + Display over apps; notification optional", "This is the fallback when Accessibility is off/unhealthy", 3,
        "UsageStats is delayed/coalesced on some OEMs; brief flashes, background start restrictions, and higher battery use are possible.",
        "Measure 100 blocked launches across Pixel/Samsung/Xiaomi/Oppo/Vivo, screen states, battery modes, rapid switching, split screen, and recents.",
        "android/.../FocusUsageMonitorService.kt; android/.../UsageStatsReader.kt; android/.../FocusBlockOverlayController.kt",
        priority="P0",
    ),
    feature(
        "F-029", "Blocking engine", "Optional Accessibility fast observer", STATUS_LIMITED,
        "When manually enabled, blocked app transitions are noticed faster and the same Focus gate is opened.",
        "Accessibility service subscribes only to TYPE_WINDOW_STATE_CHANGED and uses package names; it does not retrieve node text.",
        "Optional Accessibility service access", "Primary observer only while the service heartbeat is healthy", 4,
        "Some payment/banking apps refuse to run while any Accessibility service is enabled; Google Play policy/disclosure compliance is required.",
        "Enable service; verify package-only events and block latency; test payment/banking apps; revoke/kill service mid-session.",
        "android/.../SelfControlAccessibilityService.kt; android/res/xml/self_control_accessibility_service.xml; android/.../ProtectionAccess.kt",
        priority="P0",
    ),
    feature(
        "F-030", "Blocking engine", "Accessibility heartbeat and automatic fallback", STATUS_IMPLEMENTED,
        "If Accessibility is enabled but not actually running, compatible protection resumes automatically.",
        "Accessibility emits an in-process heartbeat every 3s; compatible service owns Focus unless the heartbeat is fresh (maximum age 10s). Daily limits remain monitored.",
        "No new permission", "Defines engine ownership", 4,
        "In-process heartbeat resets after process death; overlap/thrash during service reconnect must be stress-tested.",
        "During Focus kill/restart Accessibility service, toggle it in Settings, and kill the app process; confirm one gate and no unprotected gap.",
        "android/.../ProtectionAccess.kt; android/.../SelfControlAccessibilityService.kt; android/.../FocusUsageMonitorService.kt",
        priority="P0",
    ),
    feature(
        "F-031", "Blocking UI", "Immediate safety overlay", STATUS_LIMITED,
        "A minimal opaque block surface appears immediately while the full Focus gate Activity starts.",
        "FocusBlockOverlayController uses SYSTEM_ALERT_WINDOW from the compatible observer and is hidden only after the gate broadcasts visible state.",
        "Display over apps", "Not dependent", 4,
        "Overlay may be blocked by secure/system windows; it must never remain stuck above allowed apps.",
        "Launch blocked apps 100 times; slow device with animation scale; rotate; lock/unlock; crash gate; verify no flash and no stuck overlay.",
        "android/.../FocusBlockOverlayController.kt; android/.../FocusUsageMonitorService.kt; android/.../FocusGateActivity.kt",
        priority="P0",
    ),
    feature(
        "F-032", "Blocking UI", "Full-screen Focus gate", STATUS_IMPLEMENTED,
        "Shows app icon, blocked reason, live timer/status, break control, and only the exit action allowed by the mode.",
        "A native singleTask Activity is launched above the blocked app; it appears as a TrustIssue task/card and finishes when Focus is no longer applicable.",
        "Display over apps is used by protection; Activity itself needs none", "Same UI for both engines", 4,
        "It does not force-stop the target app; the app remains in memory/background. Recents behavior and OEM activity launch rules vary.",
        "Test block gate from launcher, recents, notification deep link, PiP, split screen, and repeated taps; inspect recents card and back behavior.",
        "android/.../FocusGateActivity.kt; android/app/src/main/AndroidManifest.xml",
        priority="P0",
    ),
    feature(
        "F-033", "Blocking UI", "Gate-closing animation", STATUS_IMPLEMENTED,
        "Block screen enters with two panels closing and a short fade/scale so the transition feels like a gate rather than a sudden popup.",
        "Native view animations run for about 180–190ms before settling into the blocking content.",
        "None", "Not dependent", 4,
        "Animation scale settings, low-end devices, and repeated Activity reuse may produce jank.",
        "Record 60fps/120fps screen video on low/mid/high devices with animator scale 0x/0.5x/1x/2x.",
        "android/.../FocusGateActivity.kt; android/.../FocusBlockOverlayController.kt",
        priority="P1",
    ),
    feature(
        "F-034", "Media", "Blocked-app audio containment", STATUS_LIMITED,
        "Opaque gate/background transition should naturally pause foreground video/audio; if audio remains active, TrustIssue briefly takes audio focus.",
        "After a 350ms settle delay, BlockedMediaController requests transient audio focus only when Android reports active music, then releases it after 850ms.",
        "No special permission", "Works in both engines", 3,
        "Android cannot identify which app owns global active audio here; unrelated Spotify/music may briefly duck/pause, but no global media-key is sent.",
        "Test YouTube/Instagram/VLC audio plus Spotify playing separately; verify blocked media stops and unrelated music resumes automatically.",
        "android/.../BlockedMediaController.kt; android/.../SelfControlAccessibilityService.kt; android/.../FocusUsageMonitorService.kt",
        priority="P0",
    ),
    feature(
        "F-035", "Daily limits", "Per-app daily limits", STATUS_IMPLEMENTED,
        "Lets the user set 15–240 minute presets (or persisted values up to 24h); blocked app receives a daily-limit gate after usage is exhausted.",
        "UsageStatsReader aggregates today's package use; the all-day compatible foreground service compares totals with persisted limits.",
        "Usage Access + Display over apps; notification optional", "Not required; compatible monitor continues even when Accessibility handles Focus", 3,
        "UsageStats accuracy and midnight/session boundaries vary; service survival depends on OEM battery policy.",
        "Set short test limits; cross midnight; change time zone; reboot; use app in split screen/PiP; compare Android usage totals.",
        "lib/screens/blocks_screen.dart; lib/screens/home_shell.dart; android/.../UsageStatsReader.kt; android/.../FocusUsageMonitorService.kt",
        priority="P0",
    ),
    feature(
        "F-036", "Daily limits", "Daily limit add/edit/remove and progress", STATUS_IMPLEMENTED,
        "Blocks tab shows configured apps, used/limit progress, and add/edit/remove controls.",
        "Flutter loads today's native usage summary and updates a locally persisted package-to-minutes map; native protection is resynchronised.",
        "Usage Access for real progress; Overlay to enforce", "Not required", 4,
        "UI can show a saved-but-not-active limit if required access is denied; the message must remain obvious.",
        "Add, edit, remove, deny permissions, reopen app, and verify UI accurately distinguishes saved versus active.",
        "lib/screens/blocks_screen.dart; lib/screens/home_shell.dart; lib/services/settings_store.dart",
    ),
    feature(
        "F-037", "Web protection", "Allowed-browser local VPN", STATUS_LIMITED,
        "An allowed browser can be used during Focus while adult, bypass, blocked-app, and custom domains are filtered.",
        "Only allowed browser packages are added to a split-tunnel VpnService; their synthetic-DNS traffic is parsed and forwarded over encrypted DNS.",
        "One-time Android VPN consent + Internet/network state; foreground notification", "Not required", 3,
        "DNS filtering cannot inspect page content or paths and cannot guarantee coverage against every in-app encrypted DNS/proxy protocol.",
        "Test Chrome/Brave/Firefox/Samsung Internet across Wi-Fi/mobile, Private DNS, browser Secure DNS, QUIC, VPN conflict, and network switching.",
        "android/.../FocusWebProtectionService.kt; android/.../DnsPacketCodec.kt; android/.../WebProtectionConfig.kt",
        priority="P0",
    ),
    feature(
        "F-038", "Web protection", "Mandatory domain categories", STATUS_LIMITED,
        "Blocks bundled adult, proxy, secure-DNS/bypass, and service-domain rules before upstream resolution.",
        "Normalized hostname and subdomain matching returns NXDOMAIN and broadcasts a category for the local explanation screen.",
        "VPN consent when browser allowed", "Not required", 3,
        "Bundled lists are intentionally small and become stale; Cloudflare Family provides broader upstream categorisation but is an external dependency.",
        "Run curated domain fixtures including subdomains, IDN/punycode, trailing dots, uppercase, malformed names, and newly registered bypass sites.",
        "android/.../WebProtectionConfig.kt; android/.../FocusWebProtectionService.kt",
        priority="P0",
    ),
    feature(
        "F-039", "Web protection", "Blocked app's web version", STATUS_LIMITED,
        "When a known service app is blocked, matching web domains are also blocked in an allowed browser.",
        "WebProtectionConfig maps selected package rules to curated service-domain sets and labels the block as app + web rule.",
        "VPN consent", "Not required", 3,
        "Only mapped services/domains are covered; mirrors, alternate domains, and web search results can bypass.",
        "For every mapped package, block the app and test canonical, mobile, short-link, alternate, and login domains.",
        "android/.../WebProtectionConfig.kt; android/.../WebBlockActivity.kt",
    ),
    feature(
        "F-040", "Web protection", "Custom website rules", STATUS_IMPLEMENTED,
        "User can add/edit/remove custom domains from app selection or Blocks.",
        "Flutter normalizes and persists a domain set; VPN policy performs exact/subdomain matching.",
        "VPN consent when browser allowed", "Not required", 4,
        "IP addresses, URL paths, search keywords, and domains contacted without DNS are outside scope.",
        "Add valid/invalid domains and subdomains; restart app/session; confirm only matching hostnames are blocked.",
        "lib/screens/select_apps_screen.dart; lib/screens/blocks_screen.dart; android/.../WebProtectionConfig.kt",
    ),
    feature(
        "F-041", "Web protection", "Encrypted family DNS and SafeSearch", STATUS_LIMITED,
        "Allowed DNS requests go to Cloudflare Family over HTTPS; selected search/video hostnames receive SafeSearch aliases.",
        "HTTPS DNS-message POST, 30-second response cache, and CNAME rewriting are implemented in the VPN service.",
        "Internet + VPN consent", "Not required", 3,
        "External resolver availability/privacy terms, certificate/network failures, captive portals, and service hostname changes affect results.",
        "Validate known adult/malware domains, SafeSearch on supported engines, cache behavior, captive portal, offline/reconnect, and TLS interception.",
        "android/.../FocusWebProtectionService.kt; android/.../DnsPacketCodec.kt",
        priority="P0",
    ),
    feature(
        "F-042", "Web protection", "Fail-closed browser protection", STATUS_IMPLEMENTED,
        "If VPN permission/start/health fails, allowed browsers are blocked instead of silently becoming unprotected.",
        "Start flow moves browsers into the block policy when consent is declined; running service transitions through starting/active/reconnecting/degraded and the observer checks browser coverage.",
        "Usage/Overlay plus conditional VPN consent", "Works with either foreground engine", 4,
        "A race between browser launch and VPN state must be stress-tested; false fail-closed blocks are safer but frustrating.",
        "Decline/revoke VPN, kill VPN service, remove network, install browser mid-session, and force upstream failure; browser must not open unprotected.",
        "lib/screens/home_shell.dart; android/.../WebProtectionConfig.kt; android/.../FocusWebProtectionService.kt",
        priority="P0",
    ),
    feature(
        "F-043", "Web protection", "Website block explanation screen", STATUS_IMPLEMENTED,
        "Shows the blocked domain/category and safe-return actions without displaying page content or history.",
        "WebBlockActivity receives only normalized domain and policy category from a local broadcast/intent.",
        "None beyond active VPN", "Accessibility may issue Back before opening it; compatible mode opens the same screen", 4,
        "Browser navigation state differs; returning may re-trigger the same DNS request.",
        "Trigger every category and both unavailable/normal states; test Back, retry, close browser, and repeated redirect loops.",
        "android/.../WebBlockActivity.kt; android/.../FocusWebProtectionService.kt",
    ),
    feature(
        "F-044", "Study tools", "PDF/Gallery tool classification", STATUS_LIMITED,
        "Recognizes installed PDF readers and gallery/system picker packages as possible study tools.",
        "Native intent queries and curated system package sets classify PDF_READER or GALLERY without reading files.",
        "Package visibility queries", "Automatic foreground-origin correlation is more reliable with Accessibility", 3,
        "OEM gallery/file-manager packages and multi-role apps can be misclassified; a browser PDF handler is excluded from default-reader eligibility.",
        "Test PDF and image flows on Pixel/Samsung/Xiaomi/Oppo/Vivo, Files by Google, Drive, Photos, Gallery, and browser PDF viewers.",
        "android/.../StudyToolResolver.kt; android/app/src/main/AndroidManifest.xml",
    ),
    feature(
        "F-045", "Study tools", "Two-minute temporary study-tool access", STATUS_LIMITED,
        "A confirmation gate can grant a study tool two minutes once per tool per session, then returns to the source study app.",
        "TrackerConfig records per-session use, schedules a timeout receiver, and FocusPill offers an early return.",
        "Foreground observer + overlay", "Transition/source detection is strongest with Accessibility; compatible mode uses UsageStats heuristics", 3,
        "Without Accessibility, Android cannot always prove whether Gallery/PDF was opened by a study app; confirmation is intentional friction.",
        "Open tool from selected and unrelated apps, use once/twice, expire/background/reboot, and verify return package and timer accounting.",
        "android/.../StudyToolGateActivity.kt; android/.../StudyToolTimeoutReceiver.kt; android/.../TrackerConfig.kt",
        priority="P0",
    ),
    feature(
        "F-046", "Study tools", "Session-long selected PDF access", STATUS_LIMITED,
        "The configured default PDF reader can be allowed as a study app for the current Focus session.",
        "StudyToolGate checks selected default reader and session eligibility, then adds a session-scoped grant whose time counts as study.",
        "Foreground observer + overlay", "More reliable with Accessibility", 3,
        "Only the selected PDF reader is eligible; switching handlers mid-session intentionally triggers a gate.",
        "Grant and deny; open same/different reader repeatedly; uninstall/disable reader; confirm grant expires when Focus ends.",
        "android/.../StudyToolGateActivity.kt; android/.../StudyToolResolver.kt; android/.../TrackerConfig.kt",
    ),
    feature(
        "F-047", "Study tools", "Gallery remains temporary", STATUS_LIMITED,
        "Gallery access is never made session-long; user confirms temporary use.",
        "ToolKind.GALLERY is marked sessionEligible=false, reducing the risk that direct Gallery distraction is counted as study.",
        "Foreground observer + overlay", "More reliable with Accessibility", 3,
        "System photo pickers and galleries may use several packages; false-positive confirmation gates remain possible.",
        "Open Gallery directly versus through selected app/system picker and verify temporary-only UI and timeout.",
        "android/.../StudyToolResolver.kt; android/.../StudyToolGateActivity.kt",
    ),
    feature(
        "F-048", "Reports", "Daily/weekly/monthly reports", STATUS_IMPLEMENTED,
        "Shows a selected day, last 7 days, or last 30 days with focus time, completion, sessions, attempts, and returns.",
        "Flutter aggregates up to 30 locally stored native daily JSON metrics and compares periods.",
        "None", "Not required", 4,
        "No cloud backup; deleted app data removes reports. Aggregation accuracy depends on event recording correctness.",
        "Create controlled sessions/events across dates; independently calculate totals and compare all periods.",
        "lib/screens/analytics_screen.dart; android/.../TrackerConfig.kt; lib/services/native_bridge.dart",
    ),
    feature(
        "F-049", "Reports", "Focus trend chart and average", STATUS_IMPLEMENTED,
        "Displays a bar chart of focus by day and average per day.",
        "Flutter derives chart height and labels from local daily metrics.",
        "None", "Not required", 5,
        "Dense 30-day labels and zero-data layouts need visual testing.",
        "Test 0, 1, 7, and 30 days with extreme values and date selection.",
        "lib/screens/analytics_screen.dart",
    ),
    feature(
        "F-050", "Reports", "Distraction attempts and outcomes", STATUS_IMPLEMENTED,
        "Shows top blocked apps, attempt counts, returned-to-focus outcomes, and ended-focus outcomes.",
        "Both observers and FocusGate record package-level attempts, handled returns, blocked duration, and emergency/normal exits into daily metrics.",
        "None", "Not required", 4,
        "Rapid repeated opens are throttled, so counts represent blocking moments rather than every OS event.",
        "Script known attempts/returns/exits; verify top-five order, totals, percentage, and no double count during gate reuse.",
        "lib/screens/analytics_screen.dart; android/.../TrackerConfig.kt; android/.../FocusGateActivity.kt",
    ),
    feature(
        "F-051", "Reports", "Focused time by study app", STATUS_IMPLEMENTED,
        "Shows where focus happened and duration per selected study app.",
        "Native timing spans are attributed by package and Flutter displays the top entries.",
        "None", "Package observation is more precise with Accessibility", 3,
        "UsageStats polling can undercount short visits; Continuous timer may not attribute all focus to an app.",
        "Use two selected apps for controlled intervals; compare per-app sum with total focus and explain Continuous remainder.",
        "lib/screens/analytics_screen.dart; android/.../TrackerConfig.kt",
    ),
    feature(
        "F-052", "Reports", "Break control metrics", STATUS_IMPLEMENTED,
        "Shows break count, allotted time, actually used time, percentage, and unused time returned.",
        "Native break start/end paths increment daily counts/durations and per-app break metrics.",
        "None", "Not required", 4,
        "Crash/reboot during break must reconcile exactly once.",
        "Take manual and expired breaks, return early, reboot mid-break, and independently sum metrics.",
        "lib/screens/analytics_screen.dart; android/.../TrackerConfig.kt",
    ),
    feature(
        "F-053", "Reports", "Motivation/insight cards", STATUS_IMPLEMENTED,
        "Generates local tips from focus change, completion, returns, and top distraction.",
        "Deterministic Flutter rules choose a badge, title, explanation, and tip from period summaries.",
        "None", "Not required", 5,
        "Insights are heuristic, not clinical or predictive.",
        "Create fixtures for no data, improved focus, high completion, high return, and top distraction branches.",
        "lib/screens/analytics_screen.dart",
    ),
    feature(
        "F-054", "Privacy", "Local-only settings and analytics", STATUS_IMPLEMENTED,
        "No account/cloud sync is required; app rules, timing, attempts, breaks, and reports remain on-device.",
        "SharedPreferences and local JSON metrics store configuration and summaries; manifest disables Android backup.",
        "None", "Accessibility metadata is package-only and does not request node content", 4,
        "VPN necessarily sends allowed-browser DNS queries to Cloudflare Family; this is external network processing, not fully local.",
        "Inspect network traffic outside active VPN and storage/backup behavior; verify no app server endpoints or analytics SDKs.",
        "android/app/src/main/AndroidManifest.xml; android/.../TrackerConfig.kt; lib/services/settings_store.dart; lib/screens/settings_screen.dart",
        priority="P0",
    ),
    feature(
        "F-055", "Diagnostics", "Redacted local debug log", STATUS_IMPLEMENTED,
        "Records protection lifecycle/health messages for troubleshooting without screen content.",
        "TrustIssueDebugLog writes bounded local diagnostic entries; callers truncate package/domain/error text.",
        "None", "Records service lifecycle only", 4,
        "Package names and blocked domains can still be sensitive metadata and need retention/redaction review.",
        "Exercise features; inspect export for passwords, typed emergency reason, full URLs, messages, notification text, and unbounded growth.",
        "android/.../TrustIssueDebugLog.kt; lib/screens/settings_screen.dart; android/.../MainActivity.kt",
        priority="P0",
    ),
    feature(
        "F-056", "Diagnostics", "Export and clear diagnostics", STATUS_IMPLEMENTED,
        "Settings can export the local diagnostic log and clear it.",
        "Flutter calls native bridge methods; Android returns a shareable/export payload and clears local log on request.",
        "System share sheet when exporting", "Not required", 4,
        "Destination apps are outside TrustIssue privacy control after the user shares.",
        "Export empty/non-empty/large log; cancel share; clear and re-export; inspect content.",
        "lib/screens/settings_screen.dart; lib/services/native_bridge.dart; android/.../MainActivity.kt",
    ),
    feature(
        "F-057", "Privacy", "Clear local reports", STATUS_IMPLEMENTED,
        "User can delete locally stored report history after confirmation.",
        "Settings invokes native analytics deletion and reloads the Reports state.",
        "None", "Not required", 4,
        "Must verify no orphaned keys, active-session corruption, or stale UI cache.",
        "Create history; clear while idle; reopen/restart; confirm zero data and current configuration remains.",
        "lib/screens/settings_screen.dart; android/.../TrackerConfig.kt; android/.../MainActivity.kt",
    ),
    feature(
        "F-058", "Settings", "Permission and protection status", STATUS_IMPLEMENTED,
        "Shows Usage Access, Overlay, Accessibility, Strict DND, VPN/web status, notification, and optional battery status with contextual actions.",
        "Flutter refreshes native access/state methods on lifecycle resume and opens the correct Android settings screen.",
        "Read-only checks; user grants access in Android Settings", "Accessibility is clearly marked optional/advanced", 4,
        "Status can be stale for a moment after OEM settings return; wording must match actual feature dependencies.",
        "Toggle every permission externally, return by Back/recents, and verify status and next action within one refresh.",
        "lib/screens/settings_screen.dart; lib/services/native_bridge.dart; android/.../MainActivity.kt",
    ),
    feature(
        "F-059", "Permissions", "Just-in-time permission flow", STATUS_IMPLEMENTED,
        "Usage/Overlay are requested only when Focus or Daily Limit is activated; DND only for Strict; VPN only when a browser is allowed.",
        "HomeShell checks each prerequisite sequentially, shows a plain-language disclosure, records a pending flow, and resumes after Settings.",
        "Feature-specific", "Never required automatically; manually enabled from Advanced Protection", 5,
        "Repeated denial and partial grants need friction testing.",
        "Fresh install: trigger Normal, Daily Limit, Strict, and allowed-browser flows separately; verify no unrelated prompt appears.",
        "lib/screens/home_shell.dart; lib/screens/settings_screen.dart",
        priority="P0",
    ),
    feature(
        "F-060", "Permissions", "Optional notification permission", STATUS_IMPLEMENTED,
        "Requests notification access for foreground-service visibility, but blocking is allowed to continue if denied.",
        "Android 13+ runtime prompt result does not gate activation; system still exposes foreground service controls.",
        "POST_NOTIFICATIONS (optional)", "Not required", 4,
        "OEM behavior and user awareness differ when foreground notifications are hidden.",
        "Deny/allow on Android 12/13/14/15/16; verify service remains alive and system disclosures are present.",
        "lib/screens/home_shell.dart; android/.../MainActivity.kt; android/.../FocusUsageMonitorService.kt",
    ),
    feature(
        "F-061", "Permissions", "Optional battery optimization exemption", STATUS_IMPLEMENTED,
        "Settings offers a manual battery reliability recommendation; onboarding does not request it.",
        "Native checks PowerManager exemption and opens app-specific battery settings only after user action.",
        "Ignore battery optimizations (optional, Settings-triggered)", "Not required", 4,
        "Not exempt by default; aggressive OEM killers may still terminate services even when unrestricted.",
        "Test optimized/unrestricted states and vendor auto-start/background-lock settings across OEMs.",
        "lib/screens/settings_screen.dart; android/.../MainActivity.kt",
    ),
    feature(
        "F-062", "Reliability", "Foreground protection service", STATUS_LIMITED,
        "Keeps compatible Focus or daily-limit monitoring alive in the background.",
        "Android specialUse foreground service runs package-only observation and displays an ongoing system notification/control.",
        "FOREGROUND_SERVICE + FOREGROUND_SERVICE_SPECIAL_USE; notification optional", "Stops owning Focus only while Accessibility heartbeat is healthy", 3,
        "Google Play special-use declaration review and OEM task killers are release risks.",
        "Run 8h soak tests under Doze, battery saver, memory pressure, app swipe-away, screen off, and OEM auto-start disabled/enabled.",
        "android/.../FocusUsageMonitorService.kt; android/app/src/main/AndroidManifest.xml",
        priority="P0",
    ),
    feature(
        "F-063", "Reliability", "Session and break timeout alarms", STATUS_IMPLEMENTED,
        "Ends fixed sessions, temporary tools, and breaks even while UI is not open.",
        "AlarmManager uses elapsed-realtime wakeups with allow-while-idle and RTC fallback for one break path; receivers sanitize and reconcile state.",
        "No exact-alarm permission", "Not required", 4,
        "Alarms are inexact and can be deferred; observer/reopen reconciliation is required for immediate correction.",
        "Test Doze, reboot, force-stop/reopen, time changes, and timeout races; verify cleanup occurs once.",
        "android/.../TrackerConfig.kt; android/.../FocusSessionTimeoutReceiver.kt; android/.../FocusBreakTimeoutReceiver.kt; android/.../StudyToolTimeoutReceiver.kt",
        priority="P0",
    ),
    feature(
        "F-064", "Reliability", "Monotonic timing and wall-clock recovery", STATUS_IMPLEMENTED,
        "Protects active durations from manual clock changes while retaining a reboot recovery anchor.",
        "elapsedRealtime drives same-boot duration math; wall-clock deadline is retained for reboot and Continuous-timer recovery.",
        "None", "Not required", 4,
        "Reboot resets elapsedRealtime; daylight/timezone/manual changes and invalid stored anchors need a device test matrix.",
        "Move clock ±24h, change timezone, reboot, and simulate stale/corrupt preferences during every mode and break.",
        "android/.../TrackerConfig.kt; android/.../StartupStateSanitizer.kt",
        priority="P0",
    ),
    feature(
        "F-065", "Reliability", "Protection restoration after reboot/update/clock change", STATUS_IMPLEMENTED,
        "Reconciles active session, DND, daily limits, and compatible service after BOOT_COMPLETED, app update, TIME_SET, or TIMEZONE_CHANGED.",
        "Manifest receiver sanitizes state, invalidates usage cache, completes expired sessions, reconciles DND, and restarts monitoring when permissions remain ready.",
        "RECEIVE_BOOT_COMPLETED; previously granted Usage/Overlay/DND", "Accessibility restarts under Android; compatible service is explicitly restarted when ready", 4,
        "Receiver is not direct-boot aware, so restoration begins after user unlock. VPN start is indirectly reconciled by the active monitor and must be verified.",
        "Reboot locked/unlocked, update APK, set time/timezone, and verify Focus/limits/DND/VPN with logs.",
        "android/.../ProtectionRestoreReceiver.kt; android/app/src/main/AndroidManifest.xml",
        priority="P0",
    ),
    feature(
        "F-066", "Reliability", "Startup state sanitizer and legacy migration", STATUS_IMPLEMENTED,
        "Repairs wrong preference types, impossible combinations, expired transient state, and legacy private data.",
        "TrustIssueApplication and restoration/timeout paths call StartupStateSanitizer before protection logic.",
        "None", "Not required", 4,
        "Only known corruptions are covered; migration must be regression-tested with real previous-version data.",
        "Install/upgrade from each supported old build and inject malformed preference fixtures; confirm safe defaults without data loss.",
        "android/.../TrustIssueApplication.kt; android/.../StartupStateSanitizer.kt",
        priority="P0",
    ),
    feature(
        "F-067", "Accessibility", "Accessibility requirement badge", STATUS_IMPLEMENTED,
        "Small badge marks features that actually need optional Accessibility; Settings explains core Focus works without it.",
        "Reusable Flutter widget is placed on Advanced Protection and locked future features.",
        "None", "Informational", 5,
        "Current implemented Focus modes do not require the badge because they have compatible fallback; future features must use it consistently.",
        "Audit every toggle/card and compare its real permission dependency with badge and disclosure.",
        "lib/widgets/accessibility_requirement_badge.dart; lib/screens/settings_screen.dart; lib/screens/blocks_screen.dart",
    ),
    feature(
        "F-068", "Short-form content", "YouTube Shorts blocker", STATUS_IMPLEMENTED,
        "Optional Accessibility detector gates the Shorts player and offers a one-Short pass.",
        "Visible YouTube labels and view IDs are inspected in memory only while the feature is enabled.",
        "Separate disclosure and toggle", "Accessibility", 4,
        "Private YouTube view hierarchies can change and require app-version regression coverage.",
        "Open a Short, grant one-Short access, swipe, and confirm the gate returns without blocking normal videos.",
        "lib/screens/blocks_screen.dart; android/app/src/main/kotlin/com/trustissue/child/SelfControlAccessibilityService.kt",
        priority="P0", pending="Run device and YouTube-version regression matrix.",
    ),
    feature(
        "F-069", "Short-form content", "Instagram Reels blocker", STATUS_IMPLEMENTED,
        "Optional Accessibility detector gates the Reels viewer and offers a one-Reel pass.",
        "Visible Instagram labels and view IDs are inspected in memory only while the feature is enabled.",
        "Separate disclosure and toggle", "Accessibility", 4,
        "Fine-grained in-app content detection is fragile and Play-policy sensitive.",
        "Open a Reel, grant one-Reel access, swipe, and confirm the gate returns without blocking Home, Explore, posts, Stories, or messages.",
        "lib/screens/blocks_screen.dart; android/app/src/main/kotlin/com/trustissue/child/InstagramReelsDetector.kt",
        priority="P0", pending="Run physical-device and Instagram-version regression matrix.",
    ),
    feature(
        "F-070", "Removed features", "YouTube Study Mode", STATUS_REMOVED,
        "Feature is intentionally absent.",
        "Removed by user decision; browser/domain protection remains separate.",
        "None", "Not applicable", None,
        "Old screenshots/copy must not advertise it.",
        "Search UI, strings, settings, and navigation for active YouTube Study Mode entry points.",
        "Current Flutter UI source",
        priority="N/A", pending="Remove any residual dead strings/assets if later found.",
    ),
    feature(
        "F-071", "Removed features", "Prevent uninstall", STATUS_REMOVED,
        "No uninstall-prevention toggle is presented or implemented.",
        "Removed by user decision; ordinary consumer apps cannot reliably prevent uninstall without device-owner/enterprise management.",
        "None", "Not applicable", None,
        "Strict/Locked can still be bypassed by uninstalling or force-stopping the app.",
        "Confirm no device-admin/device-owner/uninstall interception code or UI exists.",
        "Current Flutter/Android source",
        priority="N/A", pending="Keep removed unless an enterprise device-owner edition is planned.",
    ),
    feature(
        "F-072", "Removed features", "Block split screen and floating windows", STATUS_REMOVED,
        "No dedicated strict-protection toggles exist.",
        "Removed by user decision; general blocking still attempts to detect foreground packages.",
        "None", "Not applicable", None,
        "Multi-window, PiP, and floating-window edge cases remain platform limitations for current blocking.",
        "Verify UI absence; include split-screen/PiP in core blocker testing rather than claiming dedicated prevention.",
        "Current Flutter UI source",
        priority="N/A", pending="Keep removed; document current limitation.",
    ),
    feature(
        "F-073", "Removed features", "Block phone home screen", STATUS_REMOVED,
        "No home-screen blocking option is present.",
        "Removed from the requested product scope; launcher remains a safe navigation surface.",
        "None", "Not applicable", None,
        "Users can remain on launcher but cannot use blocked apps.",
        "Confirm no UI toggle and launcher is never blocked.",
        "Current Flutter/Android source",
        priority="N/A", pending="Keep removed.",
    ),
    feature(
        "F-074", "Code health", "Duplicate Flutter EmergencyExitScreen", STATUS_LEGACY,
        "No user-facing effect; the actual exit screen is native FocusGateActivity.",
        "The Dart EmergencyExitScreen class is defined but has no active reference; native emergency exit is launched through MethodChannel.",
        "None", "Not applicable", 2,
        "Duplicate behavior can drift and confuse future maintenance.",
        "Run reference search and confirm no route constructs EmergencyExitScreen; then remove after a normal build/test cycle.",
        "lib/screens/emergency_exit_screen.dart; android/.../FocusGateActivity.kt; android/.../MainActivity.kt",
        priority="P2", pending="Remove or deliberately wire it after compile/runtime validation.",
    ),
]


PERMISSIONS = [
    {
        "Permission / access": "Usage Access (PACKAGE_USAGE_STATS)",
        "Android type": "Special app access",
        "When requested": "Only when starting Focus or activating the first Daily App Limit",
        "Required?": "Required for compatible Focus and Daily App Limits",
        "Features": "Foreground package detection; daily usage totals/limits",
        "Data used": "Package names and change/usage times",
        "Explicitly not used": "Screen text, passwords, messages, notifications",
        "Fallback / denial": "Feature is saved but cannot activate; user can retry later",
        "Accessibility relation": "Focus can shift to Accessibility for faster events, but Daily Limits still use Usage Access",
        "Risk / policy note": "OEM UsageStats delay; explain prominent core functionality",
        "Test": "Grant/deny/revoke during Focus and Daily Limit; compare totals with Digital Wellbeing.",
    },
    {
        "Permission / access": "Display over other apps (SYSTEM_ALERT_WINDOW)",
        "Android type": "Special app access",
        "When requested": "Only when starting Focus or activating a Daily App Limit",
        "Required?": "Required by current core blocking design",
        "Features": "Immediate safety overlay; floating Focus pill",
        "Data used": "No user content; draws TrustIssue UI above apps",
        "Explicitly not used": "Reading target app content",
        "Fallback / denial": "Focus/limit cannot activate",
        "Accessibility relation": "Still required even when Accessibility is enabled",
        "Risk / policy note": "OEM restrictions and secure windows; Play disclosure",
        "Test": "Grant/revoke mid-session; verify no stuck overlay and fail-safe messaging.",
    },
    {
        "Permission / access": "POST_NOTIFICATIONS",
        "Android type": "Runtime permission on Android 13+",
        "When requested": "After core Focus/limit access is ready",
        "Required?": "Optional",
        "Features": "Visible foreground-service notifications",
        "Data used": "None",
        "Explicitly not used": "Reading or deleting other apps' notifications",
        "Fallback / denial": "Blocking continues; Android system still exposes active service controls",
        "Accessibility relation": "None",
        "Risk / policy note": "User may be less aware of background protection",
        "Test": "Deny on Android 13–16 and run an 8-hour soak.",
    },
    {
        "Permission / access": "Do Not Disturb policy access (ACCESS_NOTIFICATION_POLICY)",
        "Android type": "Special policy access",
        "When requested": "Only when starting Strict or Locked Strict",
        "Required?": "Required for Strict modes by current product rule",
        "Features": "Alarms-only DND; notification/call visual suppression",
        "Data used": "Reads and temporarily changes DND filter/policy",
        "Explicitly not used": "Notification contents",
        "Fallback / denial": "Strict does not start",
        "Accessibility relation": "None",
        "Risk / policy note": "OEM/Android-version behavior; restore pre-existing DND safely",
        "Test": "Pre-enable user DND, run/cancel/expire/reboot Strict, verify exact restoration.",
    },
    {
        "Permission / access": "Android VPN consent (BIND_VPN_SERVICE)",
        "Android type": "System VPN consent",
        "When requested": "Only when an installed browser is allowed during Focus",
        "Required?": "Conditional",
        "Features": "Browser hostname filtering",
        "Data used": "Allowed-browser DNS hostnames; DNS messages forwarded to Cloudflare Family",
        "Explicitly not used": "Page body, passwords, search text, browser history, TLS decryption",
        "Fallback / denial": "Browsers are moved to blocked policy (fail closed)",
        "Accessibility relation": "None",
        "Risk / policy note": "Conflicts with another VPN; external DNS provider; Play VPN declaration",
        "Test": "Allow one browser; decline/revoke consent; verify it never opens unprotected.",
    },
    {
        "Permission / access": "Accessibility service",
        "Android type": "Special accessibility access",
        "When requested": "Never automatically; manual Advanced Protection action",
        "Required?": "Optional for all implemented core Focus modes",
        "Features": "Faster package transition detection and stronger cross-app handoffs",
        "Data used": "TYPE_WINDOW_STATE_CHANGED package names",
        "Explicitly not used": "Accessibility node text, passwords, form values, screen contents",
        "Fallback / denial": "Usage Access compatible observer handles Focus",
        "Accessibility relation": "This is the optional engine itself",
        "Risk / policy note": "Bank/payment apps may block any enabled service; stringent Play policy",
        "Test": "Payment/banking matrix with service off/on; kill service and confirm 10s heartbeat fallback.",
    },
    {
        "Permission / access": "Battery optimization exemption",
        "Android type": "Optional OS/vendor setting",
        "When requested": "Never in onboarding; only user-initiated from Settings",
        "Required?": "Optional but useful on aggressive OEMs",
        "Features": "Improves survival of compatible monitor and daily limits",
        "Data used": "None",
        "Explicitly not used": "None",
        "Fallback / denial": "Features still run but may be killed/delayed",
        "Accessibility relation": "Helpful mainly for compatible foreground service; not a guarantee for either engine",
        "Risk / policy note": "Vendor-specific background/autostart controls remain",
        "Test": "Optimized versus unrestricted 8-hour soak on top OEMs.",
    },
    {
        "Permission / access": "RECEIVE_BOOT_COMPLETED",
        "Android type": "Install-time normal permission",
        "When requested": "No user prompt",
        "Required?": "Required for reboot restoration",
        "Features": "Restores active protection/daily limits/DND after user unlock",
        "Data used": "Local stored state only",
        "Explicitly not used": "No direct-boot data before unlock",
        "Fallback / denial": "Protection resumes when the app is next opened",
        "Accessibility relation": "Android separately manages Accessibility service restart",
        "Risk / policy note": "Receiver is directBootAware=false",
        "Test": "Reboot and wait before/after unlock; confirm restoration timing.",
    },
    {
        "Permission / access": "INTERNET + ACCESS_NETWORK_STATE",
        "Android type": "Normal permissions",
        "When requested": "No user prompt",
        "Required?": "Only materially used by web protection",
        "Features": "Encrypted family DNS and network health",
        "Data used": "DNS request/response and connectivity state",
        "Explicitly not used": "No TrustIssue account or analytics upload is implemented",
        "Fallback / denial": "VPN degrades and browsers should fail closed",
        "Accessibility relation": "None",
        "Risk / policy note": "Cloudflare external dependency",
        "Test": "Offline, captive portal, TLS intercept, Wi-Fi/mobile handoff.",
    },
    {
        "Permission / access": "FOREGROUND_SERVICE + SPECIAL_USE",
        "Android type": "Manifest/service capability",
        "When requested": "No separate prompt",
        "Required?": "Required for compatible monitoring",
        "Features": "Focus and daily limit foreground service",
        "Data used": "Package-only observations and local state",
        "Explicitly not used": "Screen content",
        "Fallback / denial": "Accessibility can handle Focus, but daily limits still need compatible monitor",
        "Accessibility relation": "Warm fallback remains available",
        "Risk / policy note": "Google Play special-use subtype review required",
        "Test": "Android 14–16 service-start compliance and Play pre-launch checks.",
    },
    {
        "Permission / access": "FOREGROUND_SERVICE_SYSTEM_EXEMPTED",
        "Android type": "Manifest/service capability for VPN",
        "When requested": "No separate prompt",
        "Required?": "Used by VpnService declaration",
        "Features": "Focus web protection foreground VPN",
        "Data used": "Allowed-browser DNS",
        "Explicitly not used": "TLS payload",
        "Fallback / denial": "Browsers fail closed",
        "Accessibility relation": "None",
        "Risk / policy note": "Validate target-SDK/Play eligibility for VPN systemExempted service type",
        "Test": "Start VPN on Android 14–16 and review Play Console policy declarations.",
    },
]


TESTS = [
    ("T-001", "Onboarding", "Fresh install requests no permission", "Fresh install / cleared data", "Complete onboarding without opening a protected feature.", "No Android permission/special-access screen appears.", "P0"),
    ("T-002", "Permissions", "Normal Focus just-in-time flow", "No access granted", "Start Normal Focus and follow prompts.", "Only Usage, Overlay, then optional notifications are requested.", "P0"),
    ("T-003", "Permissions", "Strict adds only DND", "Usage/Overlay granted; DND denied", "Start Strict.", "DND disclosure appears; Accessibility and battery are not requested.", "P0"),
    ("T-004", "Permissions", "Allowed browser adds only VPN", "Focus access ready; browser allowed", "Start Focus.", "VPN disclosure/consent appears; denial blocks browser and cancels unsafe start.", "P0"),
    ("T-005", "Compatible engine", "Normal Focus without Accessibility", "Accessibility OFF; Usage/Overlay ON", "Start Focus; open each blocked app 20 times.", "Gate appears every time with acceptable latency and no unprotected use.", "P0"),
    ("T-006", "Accessibility engine", "Fast block with Accessibility", "Accessibility ON and healthy", "Start Focus; rapidly switch to blocked apps.", "Single immediate gate; no duplicate compatible gate.", "P0"),
    ("T-007", "Hybrid engine", "Heartbeat fallback", "Active Focus with Accessibility ON", "Disable/kill Accessibility service while switching apps.", "Compatible protection takes ownership within 10s and no persistent gap occurs.", "P0"),
    ("T-008", "Compatibility", "Payment app with Accessibility OFF", "Accessibility OFF", "Open representative UPI/bank/payment apps before/during Focus when allowed.", "Apps do not complain about TrustIssue Accessibility; Focus behavior follows rules.", "P0"),
    ("T-009", "Compatibility", "Payment app with Accessibility ON", "Accessibility ON", "Open same payment matrix.", "Any third-party refusal is documented; user can disable optional Accessibility and core app continues.", "P0"),
    ("T-010", "Blocking UI", "Gate in recents", "Active Focus", "Open blocked Chrome/Instagram; open recents.", "TrustIssue gate is its own card; target app remains behind/in memory; no misleading force-close claim.", "P0"),
    ("T-011", "Blocking UI", "Animation and safety overlay handoff", "Compatible engine", "Record blocked app launch in slow motion.", "No content flash beyond accepted threshold; overlay disappears only after gate is visible.", "P0"),
    ("T-012", "Blocking UI", "No stuck overlay", "Active Focus", "Rotate, lock/unlock, crash/reopen gate, end Focus.", "Overlay/pill is removed from allowed apps and after session end.", "P0"),
    ("T-013", "Media", "Blocked foreground video stops", "YouTube/VLC playing", "Block the playing app.", "Playback becomes inaudible quickly and does not continue indefinitely behind gate.", "P0"),
    ("T-014", "Media", "Spotify is not permanently paused", "Spotify in background + another blocked media app", "Trigger gate repeatedly.", "Spotify may briefly duck/pause but resumes after transient focus release.", "P0"),
    ("T-015", "Timer", "Normal timer completion", "1-minute test configuration", "Run with screen on and off.", "Ends exactly once, removes block UI/service state, reports duration.", "P0"),
    ("T-016", "Timer", "Mode maximum validation", "Idle", "Try Normal >24h, Strict >8h, Locked >3h.", "UI/native reject over-limit values; exact maximum is accepted.", "P0"),
    ("T-017", "Stopwatch", "Count-up and normal exit", "Normal Stopwatch", "Use selected/unselected apps; end Focus.", "Elapsed time is monotonic and stop confirmation works.", "P0"),
    ("T-018", "Stopwatch", "Strict Stopwatch exit", "Strict Stopwatch", "Trigger pill/home exit.", "Protected 60s/10-word emergency exit is required.", "P0"),
    ("T-019", "Locked Strict", "Locked disabled for Stopwatch", "Stopwatch selected", "Open protection choices.", "Locked is unavailable and cannot be persisted through restart.", "P0"),
    ("T-020", "Locked Strict", "No early-exit path", "Active Locked Strict", "Try home, pill, gate, back, notification, deep link, app restart.", "No emergency/end button or successful native stop path exists.", "P0"),
    ("T-021", "Locked Strict", "Natural timeout and DND release", "Short debug/test duration or boundary harness", "Let session expire with screen off.", "Session completes once and original DND state returns.", "P0"),
    ("T-022", "Strict", "Preparation wait/skip/cancel", "Strict configured", "Exercise all three paths and background interruption.", "Only wait/skip starts; cancel/interruption remains safe.", "P1"),
    ("T-023", "Strict", "Emergency exit validation", "Active Strict", "Try 0, 9, 10 words before/after 60s.", "Exit enabled only after both requirements; attempts/cancel/success recorded correctly.", "P0"),
    ("T-024", "DND", "Existing user DND restored", "User DND already set", "Start and end Strict normally/emergency/timeout.", "TrustIssue releases only its own state and restores original filter/policy.", "P0"),
    ("T-025", "DND", "Incoming communications", "Active Strict/Locked", "Send call, emergency call, SMS, heads-up, alarm.", "Behavior matches alarms-only promise and safety expectations on each OEM.", "P0"),
    ("T-026", "Breaks", "Break earning boundaries", "Timer and Stopwatch", "Cross 30/60-minute focused thresholds.", "Correct break count/budget appears once at each threshold.", "P1"),
    ("T-027", "Breaks", "Up to three selected break apps", "Break available", "Select 1/3/4 apps and launch selected/unselected apps.", "Maximum is enforced; only selected and safety apps open.", "P0"),
    ("T-028", "Breaks", "Early return preserves unused budget", "Active break", "End after 30s; take next allowed break.", "Unused time is returned exactly once and reports match.", "P1"),
    ("T-029", "Breaks", "Expiry under Doze/reboot", "Active break", "Screen off/Doze/reboot across expiry.", "Rules restore at alarm or first reconciliation; no stuck break.", "P0"),
    ("T-030", "Tracking", "Continuous timer accounting", "Book/Continuous mode", "Spend controlled time on launcher/allowed apps and break.", "Wall-clock focus excludes break and ends at expected deadline.", "P1"),
    ("T-031", "Tracking", "Selected-app accounting", "App Study", "Alternate selected and unselected apps in measured intervals.", "Only selected/approved tool time is credited within tolerance.", "P0"),
    ("T-032", "Study tools", "Two-minute quick PDF", "App Study + default PDF", "Open PDF from selected study app; grant quick access.", "Counts study, expires once, and returns to source; second grant unavailable.", "P0"),
    ("T-033", "Study tools", "Session-long default PDF", "Eligible selected PDF reader", "Grant for session; reopen repeatedly.", "Grant lasts only current Focus and counts time.", "P1"),
    ("T-034", "Study tools", "Gallery confirmation false-positive control", "App Study", "Open Gallery directly and through study app/system picker.", "Gallery never gets silent full-session access; confirmation is clear.", "P0"),
    ("T-035", "Study tools", "PDF reader removed/changed", "Configured default PDF reader", "Uninstall/disable/change handler.", "Start flow asks for a valid new reader and does not crash.", "P1"),
    ("T-036", "Daily limits", "Basic enforcement", "Usage/Overlay ready", "Set a short limit; consume it; reopen app.", "Daily-limit gate appears consistently and Focus state remains independent.", "P0"),
    ("T-037", "Daily limits", "Midnight reset", "Limit exhausted before midnight", "Cross local midnight and time zone.", "New day's usage resets without corrupting previous report.", "P0"),
    ("T-038", "Daily limits", "All-day survival", "Limit active; no Focus", "8–24h soak with app swipe-away, screen off, battery saver.", "Monitor survives or restores; notification/system disclosure remains correct.", "P0"),
    ("T-039", "Web", "All browsers blocked means no VPN", "Blocklist includes all browsers", "Start Focus.", "VPN does not start and browsers remain blocked.", "P1"),
    ("T-040", "Web", "Adult and SafeSearch filtering", "Allowed browser + VPN", "Visit curated adult and supported search/video hosts.", "Adult domains fail; SafeSearch aliases behave as designed.", "P0"),
    ("T-041", "Web", "Blocked app web mapping", "Known app blocked; browser allowed", "Visit mapped canonical/alternate domains.", "Mapped domains block with app+web category; gaps are documented.", "P1"),
    ("T-042", "Web", "Custom domains", "VPN active", "Add exact domain; test root/subdomain/non-match.", "Root/subdomain block; unrelated suffix does not.", "P1"),
    ("T-043", "Web", "Secure DNS/bypass attempts", "VPN active", "Enable browser Secure DNS/Private DNS; test proxy/DoH endpoints.", "Known bypasses block; any unknown bypass is documented as a limitation.", "P0"),
    ("T-044", "Web", "VPN consent revoked/failure", "Active Focus with browser", "Revoke VPN, kill service, break upstream DNS.", "State becomes reconnecting/degraded and browser fails closed.", "P0"),
    ("T-045", "Web", "Network transitions", "VPN active", "Wi-Fi↔mobile, offline, captive portal, IPv4/IPv6.", "Protection recovers without leaking unfiltered browser DNS.", "P0"),
    ("T-046", "Reports", "Controlled metric reconciliation", "Known scripted sessions", "Generate starts/completions/attempts/returns/exits/breaks.", "Daily, 7-day, 30-day and per-app values equal independent calculation.", "P0"),
    ("T-047", "Reports", "Clear local reports", "History present", "Clear in Settings; restart.", "History is empty, configuration and active protection remain intact.", "P1"),
    ("T-048", "Diagnostics", "Log privacy", "All features exercised", "Export log and inspect.", "No screen text, password, message, notification body, typed emergency reason, or full URL is present.", "P0"),
    ("T-049", "Recovery", "Reboot restore", "Active Focus/Strict/limit", "Reboot, wait at lock screen, unlock.", "Expired sessions finish; DND/monitor restore safely after unlock.", "P0"),
    ("T-050", "Recovery", "App update restore", "Active modes", "Install update over app.", "State sanitizer runs; protection and DND are coherent.", "P0"),
    ("T-051", "Recovery", "Manual clock/timezone changes", "Active timer/break/limit", "Move time ±24h and change zone.", "Monotonic active duration is not extended/bypassed; daily boundary behavior is defined.", "P0"),
    ("T-052", "Permissions", "Revoke Usage or Overlay mid-session", "Active Focus", "Revoke each from Settings.", "App surfaces degraded/not-ready state and does not claim protection it cannot provide.", "P0"),
    ("T-053", "Permissions", "DND revoked mid-Strict", "Active Strict", "Revoke policy access.", "State is reconciled safely; user receives clear warning; cleanup does not alter unrelated DND.", "P0"),
    ("T-054", "UI", "Large text/small screen/accessibility semantics", "Various display settings", "Use 2x font, TalkBack, 320dp width, landscape.", "Core actions remain visible, labeled, and operable.", "P1"),
    ("T-055", "Platform", "Split screen/PiP/floating windows", "Active Focus", "Open blocked apps through each mode.", "Current behavior is measured; unsupported gaps are documented, not advertised as protected.", "P0"),
    ("T-056", "Platform", "Rapid app/recents/notification switching", "Active Focus", "100 rapid transitions and deep links.", "No crash, duplicate gates, or meaningful unprotected window.", "P0"),
    ("T-057", "OEM", "Battery optimization matrix", "Pixel/Samsung/Xiaomi/Oppo/Vivo/OnePlus", "Optimized/unrestricted 8h soak.", "Survival rate and required vendor guidance are documented.", "P0"),
    ("T-058", "Android versions", "Android 10–16 compatibility", "Representative devices/emulators", "Run start/block/end, FGS, DND, VPN, reboot suite.", "No OS-specific crash or policy violation; differences are documented.", "P0"),
    ("T-059", "Store policy", "Google Play declaration review", "Release candidate", "Review Usage Access, overlay, Accessibility, VPN, and specialUse forms/disclosures.", "Declarations match actual source/data use and user-facing disclosures.", "P0"),
    ("T-060", "Performance", "CPU/battery/memory profile", "Compatible and Accessibility engines", "8h profile idle/active and 300+ apps.", "Polling, icon cache, VPN, and reports stay within agreed budgets; no leaks.", "P0"),
]


PENDING = [
    ("P-001", "Build validation", "Pending", "No Flutter/Android compile was run because the user explicitly requested no compilation.", "Run formatter/analyzer/build in a later approved session.", "P0"),
    ("P-002", "Runtime validation", "Pending", "All implemented features are source-reviewed only; none are marked device-verified.", "Execute Test Checklist and record device/build/evidence.", "P0"),
    ("P-003", "OEM reliability", "Pending", "Compatible polling, overlays, service survival, and DND vary substantially by vendor.", "Complete Pixel/Samsung/Xiaomi/Oppo/Vivo/OnePlus matrix.", "P0"),
    ("P-004", "Payment-app friction", "Known platform issue", "Some bank/payment apps reject any enabled Accessibility service even if TrustIssue reads only package changes.", "Keep Accessibility optional, explain benefit/tradeoff, and test top payment apps per market.", "P0"),
    ("P-005", "Play policy", "Pending", "Accessibility, Usage Access, overlays, VPN, and specialUse foreground service require accurate declarations and prominent disclosures.", "Legal/policy review before release; preserve package-only/no-content behavior.", "P0"),
    ("P-006", "Locked Strict safety", "High-risk validation", "No-exit behavior has high impact if timeout, reboot, or DND cleanup fails.", "Dedicated destructive/soak test plan and recovery review before public enablement.", "P0"),
    ("P-007", "Web coverage", "Platform limitation", "DNS-only VPN cannot guarantee blocking of every encrypted DNS/proxy/IP/path-level bypass.", "Publish honest scope; expand maintained rule lists and add bypass regression corpus.", "P0"),
    ("P-008", "Web external dependency", "Pending", "Cloudflare Family availability and policy affect allowed browser DNS.", "Document provider, privacy, outage behavior, and consider configurable/provider redundancy.", "P1"),
    ("P-009", "Daily-limit precision", "Pending", "UsageStats aggregation and midnight behavior need device evidence.", "Reconcile with Digital Wellbeing across Android/OEM/timezone matrix.", "P0"),
    ("P-010", "Study-tool handoff", "Platform limitation", "Without Accessibility, Android cannot always distinguish direct Gallery/PDF opening from a study-app handoff.", "Retain small confirmation gate; tune heuristics with false-positive telemetry only if privacy-safe.", "P1"),
    ("P-011", "Media ownership", "Platform limitation", "AudioManager.isMusicActive is global, so transient focus may briefly affect unrelated music.", "Test market-leading players; keep no-media-key design; consider user setting if complaints persist.", "P1"),
    ("P-012", "Multi-window/PiP", "Known gap", "Dedicated split-screen/floating-window prevention was removed, but these modes can challenge foreground detection.", "Measure and document current coverage; do not market full prevention.", "P0"),
    ("P-013", "Shorts blocker", "Runtime validation pending", "Private YouTube view IDs and accessibility event shapes can change.", "Run real-device regression across supported YouTube versions before release.", "P0"),
    ("P-014", "Reels blocker", "Runtime validation pending", "Private Instagram view IDs and accessibility event shapes can change.", "Run real-device regression across supported Instagram versions before release.", "P0"),
    ("P-015", "Dead code", "Cleanup pending", "Flutter EmergencyExitScreen duplicates native flow and appears unreferenced.", "Remove after build/reference/runtime validation.", "P2"),
    ("P-016", "Copy consistency", "Review pending", "Settings privacy copy and current gate/Home behavior must describe the actual implementation exactly.", "Run source-to-copy privacy review; avoid saying force-close.", "P1"),
    ("P-017", "Diagnostics privacy", "Review pending", "Package/domain metadata can be sensitive even without content.", "Set retention/size limits, redact where possible, and verify export fixtures.", "P0"),
    ("P-018", "Removed: YouTube Study", "Removed", "Removed by product decision.", "Keep removed and clean residual copy/assets if found.", "N/A"),
    ("P-019", "Removed: Prevent uninstall", "Removed", "Not requested and not reliable for a normal consumer app.", "Only reconsider for device-owner enterprise edition.", "N/A"),
    ("P-020", "Removed: split/floating/home block", "Removed", "Removed by product decision.", "Document limitations and keep UI absent.", "N/A"),
]


ARCHITECTURE = [
    ("Flutter UI / HomeShell", "Configuration, tabs, just-in-time disclosures, start/stop orchestration", "MethodChannel to native Android", "Persists inactive config before native activation; restores config on failure", "lib/screens/home_shell.dart"),
    ("SettingsStore", "Local Focus policy/session/limits/domain persistence and normalization", "SharedPreferences", "Clamps mode durations and disallows Locked Stopwatch", "lib/services/settings_store.dart"),
    ("MainActivity bridge", "Permission checks/settings intents, app discovery/icons, native session commands", "Flutter MethodChannel", "Returns structured success/message maps", "android/.../MainActivity.kt"),
    ("TrackerConfig", "Authoritative native session state, timing, breaks, analytics, alarms", "SharedPreferences + AlarmManager", "Monotonic timing, wall recovery, finish-if-expired reconciliation", "android/.../TrackerConfig.kt"),
    ("FocusPolicyEngine", "Pure allow/block decision from snapshot and safety flags", "Called by both observers", "Same policy across compatible/Accessibility engines", "android/.../FocusPolicyEngine.kt"),
    ("Accessibility observer", "Fast TYPE_WINDOW_STATE_CHANGED package events", "Starts native gates/pill; 3s heartbeat", "Compatible service remains warm and takes over if heartbeat stale", "android/.../SelfControlAccessibilityService.kt"),
    ("Compatible observer", "UsageStats polling for Focus and daily limits", "300ms active loop; foreground service", "Safety overlay before Activity; retains daily-limit ownership", "android/.../FocusUsageMonitorService.kt"),
    ("ProtectionAccess", "Permission readiness and observer ownership", "Usage/Overlay checks + Accessibility heartbeat", "10s stale threshold prevents enabled-but-dead service gap", "android/.../ProtectionAccess.kt"),
    ("FocusBlockOverlayController", "Immediate minimal opaque gate while Activity starts", "SYSTEM_ALERT_WINDOW", "Hidden on ACTION_GATE_VISIBLE; OEM fallback if Activity is delayed", "android/.../FocusBlockOverlayController.kt"),
    ("FocusGateActivity", "Primary blocked/daily-limit/break/exit UI and live timer", "Native singleTask Activity", "Dismisses when session no longer applies; mode-specific exit controls", "android/.../FocusGateActivity.kt"),
    ("FocusPillController", "Draggable overlay timer and quick controls", "WindowManager overlay", "Auto-hide/collapse, saved position, visibility policy", "android/.../FocusPillController.kt"),
    ("BlockedMediaController", "Stops lingering blocked-app audio without global media keys", "Transient AudioFocus", "350ms settle then 850ms focus hold/release", "android/.../BlockedMediaController.kt"),
    ("UsageStatsReader", "Today's app usage and latest foreground package", "UsageStatsManager", "2.5s usage-total cache with invalidation", "android/.../UsageStatsReader.kt"),
    ("WebProtectionConfig", "Browser discovery, domain rules, VPN state/coverage", "PackageManager + local preferences", "Fail-closed status and 15s browser cache", "android/.../WebProtectionConfig.kt"),
    ("FocusWebProtectionService", "Allowed-browser split VPN and DNS policy", "VpnService + Cloudflare Family DoH", "Health probes, 6s reconnect grace, 30s DNS cache, degraded fail-closed state", "android/.../FocusWebProtectionService.kt"),
    ("StudyToolResolver/Gate", "PDF/Gallery classification and explicit temporary/session grants", "Package/intent queries + native Activity", "Once-per-tool quick grant and timeout/return path", "android/.../StudyToolResolver.kt; android/.../StudyToolGateActivity.kt"),
    ("StrictFocusDndController", "Apply/release only TrustIssue-owned alarms-only DND", "NotificationManager policy access", "Saves previous state and reconciles after lifecycle events", "android/.../StrictFocusDndController.kt"),
    ("Restore receivers/sanitizer", "Reboot/update/time/timeout restoration", "BroadcastReceiver + application startup", "Repairs types, expires stale state, restarts compatible monitor", "android/.../ProtectionRestoreReceiver.kt; android/.../StartupStateSanitizer.kt"),
    ("Reports", "Local summaries, charts, outcomes, break and app metrics", "Native daily JSON → Flutter aggregation", "No cloud dependency; clear/export controls", "lib/screens/analytics_screen.dart; android/.../TrackerConfig.kt"),
]


COLORS = {
    "navy": "0B1712",
    "dark": "10251B",
    "green": "7ED957",
    "lime": "B7E43C",
    "mint": "DDF6D2",
    "white": "FFFFFF",
    "light": "F3F7F2",
    "gray": "5F6D65",
    "border": "CAD7CE",
    "amber": "F4B942",
    "red": "E96B68",
    "blue": "5B9BD5",
    "purple": "9B7EDE",
}


def style_title(ws, title: str, subtitle: str, last_column: int) -> None:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_column)
    cell = ws.cell(1, 1, title)
    cell.fill = PatternFill("solid", fgColor=COLORS["navy"])
    cell.font = Font(color=COLORS["white"], size=20, bold=True)
    cell.alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 34
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=last_column)
    sub = ws.cell(2, 1, subtitle)
    sub.fill = PatternFill("solid", fgColor=COLORS["dark"])
    sub.font = Font(color="D9E6DD", size=10, italic=True)
    sub.alignment = Alignment(vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 28


def write_table_sheet(
    wb: Workbook,
    name: str,
    title: str,
    subtitle: str,
    rows: list[dict[str, object]],
    widths: dict[str, float],
    table_name: str,
) -> None:
    ws = wb.create_sheet(name)
    headers = list(rows[0].keys())
    style_title(ws, title, subtitle, len(headers))
    header_row = 4
    for col, header in enumerate(headers, 1):
        c = ws.cell(header_row, col, header)
        c.fill = PatternFill("solid", fgColor=COLORS["dark"])
        c.font = Font(color=COLORS["white"], bold=True, size=10)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = Border(bottom=Side(style="thin", color=COLORS["green"]))
    ws.row_dimensions[header_row].height = 34
    status_index = headers.index("Implementation status") + 1 if "Implementation status" in headers else None
    priority_index = headers.index("Release priority") + 1 if "Release priority" in headers else (
        headers.index("Priority") + 1 if "Priority" in headers else None
    )
    for row_idx, item in enumerate(rows, header_row + 1):
        for col_idx, header in enumerate(headers, 1):
            value = item[header]
            c = ws.cell(row_idx, col_idx, value)
            c.alignment = Alignment(vertical="top", wrap_text=True)
            c.font = Font(size=9, color="1C2922")
            c.border = Border(bottom=Side(style="hair", color=COLORS["border"]))
            c.fill = PatternFill("solid", fgColor="FFFFFF" if row_idx % 2 else COLORS["light"])
        if status_index:
            status_cell = ws.cell(row_idx, status_index)
            status_value = str(status_cell.value)
            if status_value == STATUS_IMPLEMENTED:
                status_cell.fill = PatternFill("solid", fgColor="D9EFD2")
            elif status_value == STATUS_LIMITED:
                status_cell.fill = PatternFill("solid", fgColor="FFF0C7")
            elif status_value == STATUS_UI_ONLY:
                status_cell.fill = PatternFill("solid", fgColor="E9DFFC")
            elif status_value == STATUS_REMOVED:
                status_cell.fill = PatternFill("solid", fgColor="E8E8E8")
            else:
                status_cell.fill = PatternFill("solid", fgColor="FADBD9")
            status_cell.font = Font(size=9, bold=True)
        if priority_index:
            p = ws.cell(row_idx, priority_index)
            if p.value == "P0":
                p.fill = PatternFill("solid", fgColor="F8C9C7")
                p.font = Font(size=9, bold=True, color="9C1D1D")
            elif p.value == "P1":
                p.fill = PatternFill("solid", fgColor="FFE7B0")
            elif p.value == "P2":
                p.fill = PatternFill("solid", fgColor="DCEAF7")
        ws.row_dimensions[row_idx].height = 72 if len(headers) > 10 else 55
    end_row = header_row + len(rows)
    ref = f"A{header_row}:{get_column_letter(len(headers))}{end_row}"
    table = Table(displayName=table_name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium4",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=False,
        showColumnStripes=False,
    )
    ws.add_table(table)
    ws.freeze_panes = f"A{header_row + 1}"
    ws.auto_filter.ref = ref
    ws.sheet_view.showGridLines = False
    for col_idx, header in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = widths.get(header, 18)
    ws.print_title_rows = f"1:{header_row}"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True


def build_workbook() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Executive Summary"
    ws.sheet_view.showGridLines = False
    style_title(
        ws,
        "TrustIssue — Complete Feature & Reliability Audit",
        f"Source audit date: {AUDIT_DATE.isoformat()} | Scope: Flutter UI + native Android child app | IMPORTANT: no compile or runtime test was run.",
        8,
    )

    counts = Counter(str(item["Implementation status"]) for item in FEATURES)
    scored = [int(item["Expected reliability (1-5)"]) for item in FEATURES if item["Expected reliability (1-5)"] is not None]
    p0_count = sum(1 for item in FEATURES if item["Release priority"] == "P0")

    summary_rows = [
        ("Total feature entries", len(FEATURES), "Includes implemented, limited, coming-soon, removed, and legacy items."),
        ("Implemented — runtime pending", counts[STATUS_IMPLEMENTED], "Source paths exist; must still compile and run on devices."),
        ("Implemented with platform limits", counts[STATUS_LIMITED], "Android/OEM behavior prevents a 100% guarantee."),
        ("UI-only — Coming soon", counts[STATUS_UI_ONLY], "Visible but deliberately disabled; not working protection."),
        ("Removed by decision", counts[STATUS_REMOVED], "Intentionally absent from current product."),
        ("Legacy/unreferenced", counts[STATUS_LEGACY], "Cleanup candidate, not part of the active user flow."),
        ("P0 release checks", p0_count, "Must pass before calling protection reliable."),
        ("Average expected reliability", round(sum(scored) / len(scored), 2), "Design estimate only; not a measured test result."),
        ("Runtime tests completed", 0, "User instructed not to compile; every checklist item remains NOT RUN."),
    ]

    ws["A4"] = "Outcome"
    ws["B4"] = "Value"
    ws["C4"] = "Meaning"
    for cell in ws[4][:3]:
        cell.fill = PatternFill("solid", fgColor=COLORS["dark"])
        cell.font = Font(color=COLORS["white"], bold=True)
        cell.alignment = Alignment(horizontal="center")
    for r, row in enumerate(summary_rows, 5):
        for c, value in enumerate(row, 1):
            cell = ws.cell(r, c, value)
            cell.fill = PatternFill("solid", fgColor="FFFFFF" if r % 2 else COLORS["light"])
            cell.border = Border(bottom=Side(style="hair", color=COLORS["border"]))
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        ws.cell(r, 1).font = Font(bold=True, color=COLORS["dark"])
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 70
    ws.column_dimensions["D"].width = 4
    ws.column_dimensions["E"].width = 28
    ws.column_dimensions["F"].width = 18
    ws.column_dimensions["G"].width = 18
    ws.column_dimensions["H"].width = 18

    ws["E4"] = "Reliability scale"
    ws["F4"] = "Meaning"
    for cell in ws[4][4:6]:
        cell.fill = PatternFill("solid", fgColor=COLORS["dark"])
        cell.font = Font(color=COLORS["white"], bold=True)
    reliability_legend = [
        (5, "Strong deterministic source design; normal device validation still required."),
        (4, "Strong design with lifecycle/OEM edge cases to prove."),
        (3, "Useful but inherently platform/OEM/network limited."),
        (2, "Legacy/prototype/partial."),
        (1, "UI-only; no working enforcement."),
    ]
    for r, (score, meaning) in enumerate(reliability_legend, 5):
        ws.cell(r, 5, score)
        ws.cell(r, 6, meaning)
        ws.cell(r, 5).fill = PatternFill("solid", fgColor=COLORS["mint"])
        ws.cell(r, 5).font = Font(bold=True)
        ws.cell(r, 6).alignment = Alignment(wrap_text=True)

    start = 16
    ws.merge_cells(start_row=start, start_column=1, end_row=start, end_column=8)
    ws.cell(start, 1, "Bottom line")
    ws.cell(start, 1).fill = PatternFill("solid", fgColor=COLORS["green"])
    ws.cell(start, 1).font = Font(size=14, bold=True, color=COLORS["navy"])
    ws.cell(start, 1).alignment = Alignment(vertical="center")
    ws.row_dimensions[start].height = 26
    statements = [
        "Accessibility is optional for all currently implemented core Focus modes. When it is healthy, it becomes the fast package-event engine; otherwise Usage Access + Overlay remains the compatible engine.",
        "Core blocking is not a force-stop. TrustIssue launches its own opaque gate Activity/task above the target app; the target normally remains in memory/background.",
        "The compatible engine is valuable but cannot be guaranteed equally on every Android/OEM because UsageStats timing, background Activity starts, overlays, and battery killers vary.",
        "Strict = maximum 8 hours + DND + protected emergency exit. Locked Strict = Timer only, maximum 3 hours, DND, no emergency exit.",
        "The local VPN is DNS/hostname protection for allowed browsers only. It does not decrypt web pages, but it cannot promise to stop every unknown encrypted-DNS/proxy bypass.",
        "Before release, the P0 test matrix and Google Play policy review are mandatory. The workbook does not claim any feature is runtime-verified.",
    ]
    for idx, statement in enumerate(statements, start + 1):
        ws.merge_cells(start_row=idx, start_column=1, end_row=idx, end_column=8)
        cell = ws.cell(idx, 1, f"• {statement}")
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        cell.fill = PatternFill("solid", fgColor="FFFFFF" if idx % 2 else COLORS["light"])
        cell.border = Border(bottom=Side(style="hair", color=COLORS["border"]))
        ws.row_dimensions[idx].height = 36

    chart_data_start = 27
    ws.cell(chart_data_start, 1, "Status")
    ws.cell(chart_data_start, 2, "Count")
    status_order = [STATUS_IMPLEMENTED, STATUS_LIMITED, STATUS_UI_ONLY, STATUS_REMOVED, STATUS_LEGACY]
    for i, status in enumerate(status_order, chart_data_start + 1):
        ws.cell(i, 1, status)
        ws.cell(i, 2, counts[status])
    chart = BarChart()
    chart.title = "Feature status by source audit"
    chart.y_axis.title = "Count"
    chart.height = 7
    chart.width = 14
    chart.add_data(
        Reference(ws, min_col=2, min_row=chart_data_start, max_row=chart_data_start + len(status_order)),
        titles_from_data=True,
    )
    chart.set_categories(
        Reference(ws, min_col=1, min_row=chart_data_start + 1, max_row=chart_data_start + len(status_order))
    )
    chart.legend = None
    ws.add_chart(chart, "E27")
    ws.freeze_panes = "A4"
    ws.page_setup.orientation = "landscape"

    inventory_widths = {
        "ID": 10,
        "Area": 18,
        "Feature": 30,
        "Implementation status": 28,
        "What the user sees / what it does": 42,
        "How it works internally": 50,
        "Permissions / access": 32,
        "Accessibility role": 35,
        "Expected reliability (1-5)": 14,
        "Known limitations / risks": 48,
        "How to verify / accept": 50,
        "Primary source files": 45,
        "Release priority": 12,
        "Runtime validation": 28,
        "Pending action": 42,
    }
    write_table_sheet(
        wb,
        "Feature Inventory",
        "Feature Inventory",
        "Filter by Area, status, reliability, priority, permission, or Accessibility role. Reliability is a source-design estimate, not measured evidence.",
        FEATURES,
        inventory_widths,
        "FeatureInventoryTable",
    )
    inv_ws = wb["Feature Inventory"]
    reliability_col = list(FEATURES[0].keys()).index("Expected reliability (1-5)") + 1
    reliability_letter = get_column_letter(reliability_col)
    inv_ws.conditional_formatting.add(
        f"{reliability_letter}5:{reliability_letter}{4 + len(FEATURES)}",
        ColorScaleRule(
            start_type="num", start_value=1, start_color="F8696B",
            mid_type="num", mid_value=3, mid_color="FFEB84",
            end_type="num", end_value=5, end_color="63BE7B",
        ),
    )

    write_table_sheet(
        wb,
        "Permission Matrix",
        "Permission & Access Matrix",
        "Exactly when each access is requested, why it exists, what data is used, denial behavior, and the Accessibility relationship.",
        PERMISSIONS,
        {
            "Permission / access": 34,
            "Android type": 24,
            "When requested": 40,
            "Required?": 28,
            "Features": 38,
            "Data used": 38,
            "Explicitly not used": 38,
            "Fallback / denial": 42,
            "Accessibility relation": 42,
            "Risk / policy note": 42,
            "Test": 48,
        },
        "PermissionMatrixTable",
    )

    test_rows = [
        {
            "Test ID": test_id,
            "Area": area,
            "Scenario": scenario,
            "Preconditions": preconditions,
            "Steps": steps,
            "Expected result": expected,
            "Priority": priority,
            "Status": "NOT RUN",
            "Device / Android": "",
            "Evidence / notes": "",
        }
        for test_id, area, scenario, preconditions, steps, expected, priority in TESTS
    ]
    write_table_sheet(
        wb,
        "Test Checklist",
        "Release Test Checklist",
        "No test below has been executed. Fill Device/Android, evidence, and PASS/FAIL only after running the exact build on real devices.",
        test_rows,
        {
            "Test ID": 10,
            "Area": 18,
            "Scenario": 36,
            "Preconditions": 34,
            "Steps": 48,
            "Expected result": 54,
            "Priority": 10,
            "Status": 14,
            "Device / Android": 24,
            "Evidence / notes": 46,
        },
        "TestChecklistTable",
    )

    pending_rows = [
        {
            "Item ID": item_id,
            "Topic": topic,
            "Classification": classification,
            "Current truth": truth,
            "Best next action": action,
            "Priority": priority,
        }
        for item_id, topic, classification, truth, action, priority in PENDING
    ]
    write_table_sheet(
        wb,
        "Pending & Removed",
        "Pending, Limited, Coming-soon & Removed",
        "This separates unfinished validation from deliberate product removal and unavoidable Android limitations.",
        pending_rows,
        {
            "Item ID": 10,
            "Topic": 28,
            "Classification": 24,
            "Current truth": 72,
            "Best next action": 72,
            "Priority": 12,
        },
        "PendingRemovedTable",
    )

    architecture_rows = [
        {
            "Component": component,
            "Responsibility": responsibility,
            "Interface / trigger": interface,
            "Reliability / fallback": fallback,
            "Primary source": source,
        }
        for component, responsibility, interface, fallback, source in ARCHITECTURE
    ]
    write_table_sheet(
        wb,
        "Architecture",
        "Protection Architecture",
        "The important ownership and fallback relationships behind Focus, limits, web protection, study handoffs, and restoration.",
        architecture_rows,
        {
            "Component": 34,
            "Responsibility": 56,
            "Interface / trigger": 42,
            "Reliability / fallback": 64,
            "Primary source": 52,
        },
        "ArchitectureTable",
    )

    readme = wb.create_sheet("How to Read")
    readme.sheet_view.showGridLines = False
    style_title(
        readme,
        "How to use this workbook",
        "A feature can be source-complete and still be unsafe to call reliable until its test rows pass on real target devices.",
        5,
    )
    guidance = [
        ("1", "Start at Executive Summary", "Read the bottom line and remember the audit did not compile or run the app."),
        ("2", "Filter Feature Inventory", "Use Implementation status + Release priority. P0 means release-blocking validation."),
        ("3", "Read reliability correctly", "A 4/5 means the design is strong but not verified. A 3/5 often reflects Android/OEM limits, not necessarily poor code."),
        ("4", "Use Permission Matrix for onboarding", "It is the source of truth for just-in-time prompts and for explaining Accessibility is optional."),
        ("5", "Execute Test Checklist", "Change Status only to PASS/FAIL/BLOCKED and attach device/build/video/log evidence."),
        ("6", "Do not market beyond evidence", "Do not say force-close, impossible to bypass, all websites blocked, or payment apps compatible until supported by tests."),
        ("7", "Release gate", "All P0 tests, Play policy checks, Locked Strict safety tests, and OEM soak tests should pass before production."),
    ]
    headers = ["Step", "Where", "What to do"]
    for c, h in enumerate(headers, 1):
        cell = readme.cell(4, c, h)
        cell.fill = PatternFill("solid", fgColor=COLORS["dark"])
        cell.font = Font(color=COLORS["white"], bold=True)
        cell.alignment = Alignment(horizontal="center")
    for r, row in enumerate(guidance, 5):
        for c, value in enumerate(row, 1):
            cell = readme.cell(r, c, value)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.fill = PatternFill("solid", fgColor="FFFFFF" if r % 2 else COLORS["light"])
            cell.border = Border(bottom=Side(style="hair", color=COLORS["border"]))
        readme.cell(r, 1).font = Font(bold=True, color=COLORS["dark"])
        readme.row_dimensions[r].height = 42
    readme.column_dimensions["A"].width = 10
    readme.column_dimensions["B"].width = 32
    readme.column_dimensions["C"].width = 100
    readme.freeze_panes = "A4"

    # Put the operational sheets first.
    order = [
        "Executive Summary",
        "Feature Inventory",
        "Permission Matrix",
        "Test Checklist",
        "Pending & Removed",
        "Architecture",
        "How to Read",
    ]
    wb._sheets = [wb[name] for name in order]
    wb.save(OUTPUT_PATH)

    # Re-open immediately to validate the generated OOXML package.
    checked = load_workbook(OUTPUT_PATH, read_only=False, data_only=False)
    assert checked.sheetnames == order
    assert checked["Feature Inventory"].max_row == len(FEATURES) + 4
    assert checked["Test Checklist"].max_row == len(TESTS) + 4
    assert checked["Permission Matrix"].max_row == len(PERMISSIONS) + 4
    checked.close()

    print(f"Created: {OUTPUT_PATH}")
    print(f"Feature entries: {len(FEATURES)}")
    print(f"Permission rows: {len(PERMISSIONS)}")
    print(f"Test cases: {len(TESTS)}")
    print(f"Pending/removed rows: {len(PENDING)}")


if __name__ == "__main__":
    build_workbook()
